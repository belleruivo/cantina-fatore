'''
=======================================================================================================================================
Herança
Herança é um mecanismo onde uma classe (filha) herda atributos e métodos de outra classe (mãe). No seu código:

python
class RelatorioVendas(BaseRelatorio):
    def __init__(self):
        super().__init__()
Aqui, RelatorioVendas herda de BaseRelatorio, utilizando o método super() para chamar o construtor da classe mãe e garantir que self.download_dir seja inicializado. Isso permite que RelatorioVendas use os métodos save_to_excel e format_excel definidos em BaseRelatorio.

=======================================================================================================================================
Princípio Aberto/Fechado (OCP)
OCP diz que uma classe deve ser aberta para extensão, mas fechada para modificação. Seu código segue esse princípio porque:

BaseRelatorio pode ser estendida por outras classes, como RelatorioVendas, sem precisar alterar BaseRelatorio.

Novas funcionalidades podem ser adicionadas em subclasses sem modificar a classe base.

python
class RelatorioFuncionarios(BaseRelatorio):
    # nova funcionalidade específica para relatórios de funcionários

=======================================================================================================================================
Princípio da Responsabilidade Única (SRP)
SRP afirma que uma classe deve ter uma única responsabilidade. No seu código:

BaseRelatorio é responsável apenas por salvar e formatar relatórios em Excel.

RelatorioVendas é responsável apenas por buscar dados de vendas e gerar o relatório específico de vendas.

python
class BaseRelatorio:
    # funções de salvar e formatar relatórios

class RelatorioVendas(BaseRelatorio):
    # funções de buscar dados e gerar relatórios de vendas

=======================================================================================================================================
Princípio da Substituição de Liskov (LSP)
LSP afirma que uma classe filha deve ser substituível pela classe mãe sem alterar a funcionalidade do programa. No seu código:

RelatorioVendas pode substituir BaseRelatorio porque herda e utiliza os métodos de BaseRelatorio sem alterar a lógica fundamental.

    python
        def processar_relatorio(relatorio: BaseRelatorio):
            relatorio.save_to_excel(...)

        relatorio_vendas = RelatorioVendas()
        processar_relatorio(relatorio_vendas)  # funciona perfeitamente com a classe filha
'''

import os
from flask import render_template, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from app.utils.database import get_db_connection

class BaseRelatorio:
    def __init__(self):
        self.download_dir = os.path.abspath(os.path.join(os.getcwd(), 'download_relatorio'))
        if not os.path.exists(self.download_dir):
            raise Exception(f"Diretório de download não existe: {self.download_dir}")
    
    def save_to_excel(self, df, file_name, sheet_name):
        file_path = os.path.join(self.download_dir, file_name)
        if os.path.exists(file_path):
            os.remove(file_path)
        
        df.to_excel(file_path, index=False, sheet_name=sheet_name)
        
        if not os.path.exists(file_path):
            raise Exception(f"Erro ao salvar o arquivo: {file_path}")
        
        return file_path

    def format_excel(self, file_path):
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value) or ""))
                except:
                    pass
            worksheet.column_dimensions[column].width = max_length + 2

        workbook.save(file_path)

class RelatorioVendas(BaseRelatorio):
    def __init__(self):
        super().__init__()
    
    def fetch_sales_data(self):
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''SELECT v.id, v.comprador_tipo, f.nome AS comprador_nome, v.metodo_pagamento, v.total, v.data_venda
                          FROM vendas v
                          LEFT JOIN funcionarios f ON v.comprador_id = f.id''')
        vendas = cursor.fetchall()
        conn.close()
        return vendas

    def gerar_relatorio_vendas(self):
        vendas = self.fetch_sales_data()
        df = pd.DataFrame(vendas, columns=["ID", "Comprador Tipo", "Comprador Nome", "Método Pagamento", "Total", "Data Venda"])
        file_path = self.save_to_excel(df, 'relatorio_vendas.xlsx', 'Vendas')
        self.format_excel(file_path)
        return render_template("reports.html", show_sidebar=True)

    def download_relatorio(self):
        file_path = os.path.join(self.download_dir, 'relatorio_vendas.xlsx')
        return send_file(file_path, as_attachment=True, download_name="relatorio_vendas.xlsx")
